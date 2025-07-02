from flask import Blueprint, render_template, request, make_response, flash, redirect, url_for, send_file, current_app
from flask_login import login_required, current_user
from sqlalchemy.orm import Session
from sqlalchemy import func, desc, case
from config.database import db_config
from models import Pedido, PedidoProducto, Producto, Cliente
from datetime import datetime, date, timedelta
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

reportes_bp = Blueprint('reportes', __name__, url_prefix='/reportes')

def get_current_year():
    """Obtener el año actual para el footer"""
    return datetime.now().year

@reportes_bp.route('/pedidos')
def reporte_pedidos():
    """Reporte de pedidos con filtros y paginación"""
    db = db_config.get_session()
    try:
        # Parámetros de filtro
        fecha_desde = request.args.get('fecha_desde', '')
        fecha_hasta = request.args.get('fecha_hasta', '')
        estado = request.args.get('estado', '')
        cliente_id = request.args.get('cliente_id', '')
        page = int(request.args.get('page', 1))
        per_page = 20

        # Query base
        query = db.query(Pedido)

        # Aplicar filtros
        if fecha_desde:
            try:
                fecha_desde_dt = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
                query = query.filter(func.date(Pedido.fecha_creacion) >= fecha_desde_dt)
            except ValueError:
                flash('Fecha desde inválida', 'error')

        if fecha_hasta:
            try:
                fecha_hasta_dt = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
                query = query.filter(func.date(Pedido.fecha_creacion) <= fecha_hasta_dt)
            except ValueError:
                flash('Fecha hasta inválida', 'error')

        if estado:
            query = query.filter(Pedido.estado_pedido_general == estado)

        if cliente_id and cliente_id != 'todos':
            try:
                cliente_id_int = int(cliente_id)
                query = query.filter(Pedido.cliente_id == cliente_id_int)
            except ValueError:
                flash('ID de cliente inválido', 'error')

        # Ordenar por fecha de creación descendente
        query = query.order_by(desc(Pedido.fecha_creacion))

        # Paginación
        total = query.count()
        pedidos = query.offset((page - 1) * per_page).limit(per_page).all()

        # Crear objeto de paginación manual
        class Pagination:
            def __init__(self, page, per_page, total):
                self.page = page
                self.per_page = per_page
                self.total = total
                self.pages = (total + per_page - 1) // per_page
                self.has_prev = page > 1
                self.prev_num = page - 1 if self.has_prev else None
                self.has_next = page < self.pages
                self.next_num = page + 1 if self.has_next else None

            def iter_pages(self, left_edge=2, right_edge=2, left_current=2, right_current=3):
                last = self.pages
                for num in range(1, last + 1):
                    if num <= left_edge or \
                       (self.page - left_current - 1 < num < self.page + right_current) or \
                       num > last - right_edge:
                        yield num

        pagination = Pagination(page, per_page, total)

        # Obtener datos para filtros
        estados = db.query(Pedido.estado_pedido_general).distinct().filter(
            Pedido.estado_pedido_general.isnot(None)
        ).all()
        estados = [e[0] for e in estados if e[0]]

        clientes = db.query(Cliente).order_by(Cliente.nombre_comercial).all()

        current_year = get_current_year()
        
        return render_template('reporte_pedidos.html', 
                             pedidos=pedidos,
                             pagination=pagination,
                             estados=estados,
                             clientes=clientes,
                             current_user=current_user,
                             current_year=current_year)
    
    except Exception as e:
        flash(f'Error al generar el reporte: {str(e)}', 'error')
        return redirect(url_for('pedidos.lista'))
    finally:
        db.close()

@reportes_bp.route('/consolidado')
@reportes_bp.route('/consolidado-productos')  # Agregar ruta alternativa con guiones
def consolidado_productos():
    """Consolidado de productos pedidos agrupados por categoría con subtotales por formulación y referencia"""
    db = db_config.get_session()
    try:
        # Parámetros de filtro
        estado = request.args.get('estado', '')
        fecha_desde = request.args.get('fecha_desde', '')
        fecha_hasta = request.args.get('fecha_hasta', '')
        
        # Obtener parámetros de filtro
        categoria = request.args.get('categoria', '')
        formulacion = request.args.get('formulacion', '')
        
        # Si se pide reset o es primera carga, limpiar filtros
        if request.args.get('reset') or not request.args:
            categoria = ''
            formulacion = ''

        # Query base para obtener items de pedidos con información de productos e incluir presentaciones
        # Agrupa directamente por producto y suma las cantidades
        # Calcula el peso dinámicamente: usa peso_total_g_item si no es NULL, sino calcula cantidad * gramaje_g
        query = db.query(
            func.sum(PedidoProducto.cantidad).label('cantidad_total'),
            func.sum(
                case(
                    (PedidoProducto.peso_total_g_item.isnot(None), PedidoProducto.peso_total_g_item),
                    else_=PedidoProducto.cantidad * Producto.gramaje_g
                )
            ).label('peso_total'),
            PedidoProducto.comentarios_item.label('presentacion'),
            Producto.referencia_de_producto,
            Producto.formulacion_grupo,
            Producto.categoria_linea
        ).join(
            Producto, PedidoProducto.producto_id == Producto.id
        ).join(
            Pedido, PedidoProducto.pedido_id == Pedido.id
        ).group_by(
            Producto.categoria_linea,
            Producto.formulacion_grupo,
            Producto.referencia_de_producto,
            PedidoProducto.comentarios_item
        )

        # Aplicar filtros
        if estado:
            query = query.filter(Pedido.estado_pedido_general == estado)

        if fecha_desde:
            try:
                fecha_desde_dt = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
                query = query.filter(func.date(Pedido.fecha_creacion) >= fecha_desde_dt)
            except ValueError:
                flash('Fecha desde inválida', 'error')

        if fecha_hasta:
            try:
                fecha_hasta_dt = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
                query = query.filter(func.date(Pedido.fecha_creacion) <= fecha_hasta_dt)
            except ValueError:
                flash('Fecha hasta inválida', 'error')

        if categoria:
            query = query.filter(Producto.categoria_linea.ilike(f'%{categoria}%'))

        if formulacion:
            query = query.filter(Producto.formulacion_grupo.ilike(f'%{formulacion}%'))

        # Obtener resultados
        resultados = query.all()

        # Agrupar productos consolidados y calcular subtotales jerárquicos
        productos_consolidados = []
        subtotales_formulacion = {}
        subtotales_categoria = {}
        total_cantidad = 0
        total_peso = 0

        for resultado in resultados:
            categoria = resultado.categoria_linea or 'Sin Categoría'
            formulacion = resultado.formulacion_grupo or 'Sin Formulación'
            referencia = resultado.referencia_de_producto or 'Sin Referencia'
            presentacion = resultado.presentacion or 'Sin Presentación'
            cantidad = resultado.cantidad_total or 0
            peso = resultado.peso_total or 0
            
            producto_consolidado = {
                'categoria': categoria,
                'formulacion': formulacion,
                'referencia': referencia,
                'presentacion': presentacion,
                'total_cantidad': cantidad,
                'total_peso': peso
            }
            
            productos_consolidados.append(producto_consolidado)
            
            # Calcular subtotales por formulación (categoria + formulacion)
            form_key = f"{categoria}|{formulacion}"
            if form_key not in subtotales_formulacion:
                subtotales_formulacion[form_key] = {'cantidad': 0, 'peso': 0}
            subtotales_formulacion[form_key]['cantidad'] += cantidad
            subtotales_formulacion[form_key]['peso'] += peso
            
            # Calcular subtotales por categoría
            if categoria not in subtotales_categoria:
                subtotales_categoria[categoria] = {'cantidad': 0, 'peso': 0}
            subtotales_categoria[categoria]['cantidad'] += cantidad
            subtotales_categoria[categoria]['peso'] += peso
            
            # Actualizar totales generales
            total_cantidad += cantidad
            total_peso += peso

        # Ordenar productos por categoría, formulación, referencia, presentación
        productos_consolidados.sort(key=lambda x: (x['categoria'], x['formulacion'], x['referencia'], x['presentacion']))

        # Obtener datos para filtros (independientes de los filtros aplicados)
        estados = db.query(Pedido.estado_pedido_general).distinct().filter(
            Pedido.estado_pedido_general.isnot(None)
        ).all()
        estados = [e[0] for e in estados if e[0]]

        # Obtener todas las categorías y formulaciones disponibles sin filtros
        categorias = db.query(Producto.categoria_linea).distinct().filter(
            Producto.categoria_linea.isnot(None)
        ).all()
        categorias = sorted([c[0] for c in categorias if c[0]])

        formulaciones = db.query(Producto.formulacion_grupo).distinct().filter(
            Producto.formulacion_grupo.isnot(None)
        ).all()
        formulaciones = sorted([f[0] for f in formulaciones if f[0]])

        # Crear objeto de filtros para pasar a la plantilla
        filtros = {
            'estado': estado,
            'fecha_desde': fecha_desde,
            'fecha_hasta': fecha_hasta,
            'categoria': categoria,
            'formulacion': formulacion
        }

        current_year = get_current_year()

        return render_template('consolidado_productos.html',
                             productos_consolidados=productos_consolidados,
                             subtotales_formulacion=subtotales_formulacion,
                             subtotales_categoria=subtotales_categoria,
                             total_cantidad=total_cantidad,
                             total_peso=total_peso,
                             estados=estados,
                             categorias=categorias,
                             formulaciones=formulaciones,
                             filtros=filtros,
                             current_user=current_user,
                             current_year=current_year)

    except Exception as e:
        flash(f'Error al generar el consolidado: {str(e)}', 'error')
        return redirect(url_for('pedidos.lista'))
    finally:
        db.close()

@reportes_bp.route('/exportar-pedidos-excel')
def exportar_pedidos_excel():
    """Exportar reporte de pedidos a Excel con estructura jerárquica por cliente"""
    db = db_config.get_session()
    try:
        # Aplicar los mismos filtros que en el reporte
        fecha_desde = request.args.get('fecha_desde', '')
        fecha_hasta = request.args.get('fecha_hasta', '')
        estado = request.args.get('estado', '')
        cliente_id = request.args.get('cliente_id', '')

        # Query base con join para obtener información completa
        query = db.query(Pedido).join(
            PedidoProducto, Pedido.id == PedidoProducto.pedido_id
        ).join(
            Producto, PedidoProducto.producto_id == Producto.id
        ).outerjoin(
            Cliente, Pedido.cliente_id == Cliente.id
        )

        # Aplicar filtros
        if fecha_desde:
            try:
                fecha_desde_dt = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
                query = query.filter(func.date(Pedido.fecha_creacion) >= fecha_desde_dt)
            except ValueError:
                pass

        if fecha_hasta:
            try:
                fecha_hasta_dt = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
                query = query.filter(func.date(Pedido.fecha_creacion) <= fecha_hasta_dt)
            except ValueError:
                pass

        if estado:
            query = query.filter(Pedido.estado_pedido_general == estado)

        if cliente_id and cliente_id != 'todos':
            try:
                cliente_id_int = int(cliente_id)
                query = query.filter(Pedido.cliente_id == cliente_id_int)
            except ValueError:
                pass

        # Obtener pedidos únicos y ordenarlos
        pedidos = query.distinct().order_by(desc(Pedido.fecha_creacion)).all()

        # Agrupar pedidos por cliente
        pedidos_por_cliente = {}
        for pedido in pedidos:
            cliente_key = (
                pedido.cliente_asociado.nombre_comercial if pedido.cliente_asociado 
                else pedido.nombre_cliente_ingresado or 'Cliente no registrado'
            )
            
            if cliente_key not in pedidos_por_cliente:
                pedidos_por_cliente[cliente_key] = []
            
            pedidos_por_cliente[cliente_key].append(pedido)

        # Crear libro de Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte de Pedidos"

        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cliente_font = Font(bold=True, color="FFFFFF", size=14)
        cliente_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")  # Azul
        pedido_font = Font(bold=True, color="000000")
        pedido_fill = PatternFill(start_color="B3D9FF", end_color="B3D9FF", fill_type="solid")  # Azul claro
        center_alignment = Alignment(horizontal="center")
        left_alignment = Alignment(horizontal="left")

        # Encabezados
        headers = [
            "Código Producto", "Referencia", "Cantidad", "Comentarios", "Dirección", "Horarios"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment

        # Datos con estructura jerárquica
        row = 2
        
        if pedidos_por_cliente:
            for cliente_nombre, pedidos_cliente in pedidos_por_cliente.items():
                # Nivel 1: Cliente (azul, texto grande)
                cliente_cell = ws.cell(row=row, column=1, value=f"👤 {cliente_nombre}")
                cliente_cell.font = cliente_font
                cliente_cell.fill = cliente_fill
                cliente_cell.alignment = left_alignment
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
                row += 1
                
                for pedido in pedidos_cliente:
                    # Nivel 2: Productos del Pedido dentro de Cliente (azul claro)
                    pedido_info = f"📦 Pedido #{pedido.id} - {pedido.fecha_creacion.strftime('%d/%m/%Y')} - Estado: {pedido.estado_pedido_general}"
                    pedido_cell = ws.cell(row=row, column=1, value=pedido_info)
                    pedido_cell.font = pedido_font
                    pedido_cell.fill = pedido_fill
                    pedido_cell.alignment = left_alignment
                    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
                    row += 1
                    
                    # Consolidar productos iguales dentro del mismo pedido
                    productos_consolidados = {}
                    
                    for item in pedido.items:
                        if item.producto_asociado:
                            producto_key = (item.producto_asociado.codigo, item.producto_asociado.referencia_de_producto)
                            
                            if producto_key not in productos_consolidados:
                                productos_consolidados[producto_key] = {
                                    'codigo': item.producto_asociado.codigo,
                                    'referencia': item.producto_asociado.referencia_de_producto,
                                    'cantidad_total': 0,
                                    'presentaciones': []
                                }
                            
                            # Sumar cantidad
                            productos_consolidados[producto_key]['cantidad_total'] += item.cantidad or 0
                            
                            # Recopilar presentaciones únicas
                            if item.comentarios_item and item.comentarios_item.strip():
                                if item.comentarios_item not in productos_consolidados[producto_key]['presentaciones']:
                                    productos_consolidados[producto_key]['presentaciones'].append(item.comentarios_item)
                    
                    # Nivel 3: Items consolidados de productos
                    for producto_key, producto_data in productos_consolidados.items():
                        # Código del producto
                        ws.cell(row=row, column=1, value=producto_data['codigo'])
                        
                        # Referencia del producto
                        ws.cell(row=row, column=2, value=producto_data['referencia'])
                        
                        # Cantidad total consolidada
                        ws.cell(row=row, column=3, value=f"{producto_data['cantidad_total']} unidades")
                        
                        # Presentaciones consolidadas
                        presentaciones_texto = "; ".join(producto_data['presentaciones']) if producto_data['presentaciones'] else 'Sin presentaciones'
                        ws.cell(row=row, column=4, value=presentaciones_texto)
                        
                        # Dirección
                        direccion_completa = []
                        if pedido.direccion_entrega:
                            direccion_completa.append(pedido.direccion_entrega)
                        if pedido.ciudad_entrega:
                            direccion_completa.append(pedido.ciudad_entrega)
                        if pedido.departamento_entrega:
                            direccion_completa.append(pedido.departamento_entrega)
                        
                        direccion = ", ".join(direccion_completa) if direccion_completa else 'Sin dirección especificada'
                        ws.cell(row=row, column=5, value=direccion)
                        
                        # Horarios
                        horario = pedido.despacho_horario_atencion if pedido.despacho_horario_atencion else 'Sin horario especificado'
                        ws.cell(row=row, column=6, value=horario)
                        
                        row += 1
                    
                    # Línea en blanco para separar pedidos
                    row += 1
                
                # Línea en blanco adicional para separar clientes
                row += 1
        else:
            # No hay resultados
            no_results_cell = ws.cell(row=row, column=1, value="No se encontraron pedidos con los filtros seleccionados")
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            no_results_cell.alignment = center_alignment

        # Ajustar ancho de columnas
        column_widths = [18, 35, 15, 30, 40, 25]  # Anchos específicos para cada columna
        for i, width in enumerate(column_widths, 1):
            column_letter = openpyxl.utils.get_column_letter(i)
            ws.column_dimensions[column_letter].width = width

        # Crear respuesta
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=reporte_pedidos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return response

    except Exception as e:
        flash(f'Error al exportar: {str(e)}', 'error')
        return redirect(url_for('reportes.reporte_pedidos'))
    finally:
        db.close()

@reportes_bp.route('/exportar-consolidado-excel')
def exportar_consolidado_excel():
    """Exportar consolidado de productos a Excel agrupado por categoría con subtotales por formulación"""
    db = db_config.get_session()
    try:
        # Aplicar los mismos filtros que en el consolidado
        estado = request.args.get('estado', '')
        fecha_desde = request.args.get('fecha_desde', '')
        fecha_hasta = request.args.get('fecha_hasta', '')
        categoria = request.args.get('categoria', '')
        formulacion = request.args.get('formulacion', '')

        # Query base (misma lógica que en consolidado_productos)
        # Agrupa directamente por producto y suma las cantidades
        # Calcula el peso dinámicamente: usa peso_total_g_item si no es NULL, sino calcula cantidad * gramaje_g
        query = db.query(
            func.sum(PedidoProducto.cantidad).label('cantidad_total'),
            func.sum(
                case(
                    (PedidoProducto.peso_total_g_item.isnot(None), PedidoProducto.peso_total_g_item),
                    else_=PedidoProducto.cantidad * Producto.gramaje_g
                )
            ).label('peso_total'),
            PedidoProducto.comentarios_item.label('presentacion'),
            Producto.referencia_de_producto,
            Producto.formulacion_grupo,
            Producto.categoria_linea
        ).join(
            Producto, PedidoProducto.producto_id == Producto.id
        ).join(
            Pedido, PedidoProducto.pedido_id == Pedido.id
        ).group_by(
            Producto.categoria_linea,
            Producto.formulacion_grupo,
            Producto.referencia_de_producto,
            PedidoProducto.comentarios_item
        )

        # Aplicar filtros
        if estado:
            query = query.filter(Pedido.estado_pedido_general == estado)
        if fecha_desde:
            try:
                fecha_desde_dt = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
                query = query.filter(func.date(Pedido.fecha_creacion) >= fecha_desde_dt)
            except ValueError:
                pass
        if fecha_hasta:
            try:
                fecha_hasta_dt = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
                query = query.filter(func.date(Pedido.fecha_creacion) <= fecha_hasta_dt)
            except ValueError:
                pass
        if categoria:
            query = query.filter(Producto.categoria_linea.ilike(f'%{categoria}%'))
        if formulacion:
            query = query.filter(Producto.formulacion_grupo.ilike(f'%{formulacion}%'))

        resultados = query.all()

        # Procesar productos consolidados y calcular subtotales jerárquicos para Excel
        productos_consolidados = []
        subtotales_formulacion_excel = {}
        subtotales_categoria_excel = {}
        total_cantidad = 0
        total_peso = 0

        for resultado in resultados:
            categoria = resultado.categoria_linea or 'Sin Categoría'
            formulacion = resultado.formulacion_grupo or 'Sin Formulación'
            referencia = resultado.referencia_de_producto or 'Sin Referencia'
            presentacion = resultado.presentacion or 'Sin Presentación'
            cantidad = resultado.cantidad_total or 0
            peso = resultado.peso_total or 0
            
            producto_consolidado = {
                'categoria': categoria,
                'formulacion': formulacion,
                'referencia': referencia,
                'presentacion': presentacion,
                'total_cantidad': cantidad,
                'total_peso': peso
            }
            
            productos_consolidados.append(producto_consolidado)
            
            # Calcular subtotales por formulación
            form_key = f"{categoria}|{formulacion}"
            if form_key not in subtotales_formulacion_excel:
                subtotales_formulacion_excel[form_key] = {'cantidad': 0, 'peso': 0}
            subtotales_formulacion_excel[form_key]['cantidad'] += cantidad
            subtotales_formulacion_excel[form_key]['peso'] += peso
            
            # Calcular subtotales por categoría
            if categoria not in subtotales_categoria_excel:
                subtotales_categoria_excel[categoria] = {'cantidad': 0, 'peso': 0}
            subtotales_categoria_excel[categoria]['cantidad'] += cantidad
            subtotales_categoria_excel[categoria]['peso'] += peso
            
            # Actualizar totales generales
            total_cantidad += cantidad
            total_peso += peso

        # Ordenar productos por categoría, formulación, referencia, presentación
        productos_consolidados.sort(key=lambda x: (x['categoria'], x['formulacion'], x['referencia'], x['presentacion']))

        # Crear libro de Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Consolidado de Productos"

        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        categoria_font = Font(bold=True, color="000000", size=14)
        categoria_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")  # Azul
        formulacion_font = Font(bold=True, color="000000")
        formulacion_fill = PatternFill(start_color="D1ECF1", end_color="D1ECF1", fill_type="solid")  # Azul claro
        subtotal_font = Font(bold=True, color="000000")
        subtotal_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")  # Amarillo claro
        total_categoria_font = Font(bold=True, color="000000")
        total_categoria_fill = PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid")  # Gris claro
        total_font = Font(bold=True, color="FFFFFF")
        total_fill = PatternFill(start_color="343A40", end_color="343A40", fill_type="solid")  # Gris oscuro
        center_alignment = Alignment(horizontal="center")
        right_alignment = Alignment(horizontal="right")

        # Encabezados
        headers = [
            "Categoría", "Formulación", "Referencia de Producto", "Presentación",
            "Cantidad Total", "Peso Total (g)"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment

        # Datos consolidados con subtotales jerárquicos (sin encabezados intermedios)
        row = 2
        current_categoria = ''
        current_formulacion = ''
        current_referencia = ''
        
        if productos_consolidados:
            for i, producto in enumerate(productos_consolidados):
                # Actualizar variables de control sin mostrar encabezados
                if producto['categoria'] != current_categoria:
                    current_categoria = producto['categoria']
                    current_formulacion = ''
                    current_referencia = ''
                
                if producto['formulacion'] != current_formulacion:
                    current_formulacion = producto['formulacion']
                    current_referencia = ''
                
                if producto['referencia'] != current_referencia:
                    current_referencia = producto['referencia']
                
                # Fila del producto
                ws.cell(row=row, column=1, value=producto['categoria'])
                ws.cell(row=row, column=2, value=producto['formulacion'])
                ws.cell(row=row, column=3, value=producto['referencia'])
                ws.cell(row=row, column=4, value=producto['presentacion'])
                
                cantidad_cell = ws.cell(row=row, column=5, value=round(producto['total_cantidad'], 2))
                cantidad_cell.alignment = right_alignment
                peso_cell = ws.cell(row=row, column=6, value=round(producto['total_peso'], 2))
                peso_cell.alignment = right_alignment
                
                row += 1
                
                # Verificar si es el último producto del grupo para mostrar subtotales
                next_producto = productos_consolidados[i + 1] if i + 1 < len(productos_consolidados) else None
                
                # Subtotal por Formulación
                if not next_producto or next_producto['categoria'] != producto['categoria'] or next_producto['formulacion'] != producto['formulacion']:
                    form_key = f"{producto['categoria']}|{producto['formulacion']}"
                    
                    cell = ws.cell(row=row, column=4, value=f"🧮 Subtotal {producto['formulacion']}:")
                    cell.font = subtotal_font
                    cell.fill = subtotal_fill
                    cell.alignment = right_alignment
                    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
                    
                    cantidad_cell = ws.cell(row=row, column=5, value=round(subtotales_formulacion_excel[form_key]['cantidad'], 2))
                    cantidad_cell.font = subtotal_font
                    cantidad_cell.fill = subtotal_fill
                    cantidad_cell.alignment = right_alignment
                    
                    peso_cell = ws.cell(row=row, column=6, value=round(subtotales_formulacion_excel[form_key]['peso'], 2))
                    peso_cell.font = subtotal_font
                    peso_cell.fill = subtotal_fill
                    peso_cell.alignment = right_alignment
                    
                    row += 1
                
                # Subtotal por Categoría
                if not next_producto or next_producto['categoria'] != producto['categoria']:
                    cell = ws.cell(row=row, column=4, value=f"🏷️ Total {producto['categoria']}:")
                    cell.font = total_categoria_font
                    cell.fill = total_categoria_fill
                    cell.alignment = right_alignment
                    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
                    
                    cantidad_cell = ws.cell(row=row, column=5, value=round(subtotales_categoria_excel[producto['categoria']]['cantidad'], 2))
                    cantidad_cell.font = total_categoria_font
                    cantidad_cell.fill = total_categoria_fill
                    cantidad_cell.alignment = right_alignment
                    
                    peso_cell = ws.cell(row=row, column=6, value=round(subtotales_categoria_excel[producto['categoria']]['peso'], 2))
                    peso_cell.font = total_categoria_font
                    peso_cell.fill = total_categoria_fill
                    peso_cell.alignment = right_alignment
                    
                    row += 1
                    # Línea en blanco para separar categorías
                    row += 1
        else:
            # No hay resultados
            no_results_cell = ws.cell(row=row, column=1, value="No se encontraron resultados con los filtros seleccionados")
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            no_results_cell.alignment = center_alignment
            row += 1

        # Total General
        total_cell = ws.cell(row=row, column=4, value="🔢 TOTALES GENERALES:")
        total_cell.font = total_font
        total_cell.fill = total_fill
        total_cell.alignment = right_alignment
        
        # Merge las primeras 4 columnas para el texto del total
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        
        # Total cantidad
        total_cantidad_cell = ws.cell(row=row, column=5, value=round(total_cantidad, 2))
        total_cantidad_cell.font = total_font
        total_cantidad_cell.fill = total_fill
        total_cantidad_cell.alignment = right_alignment
        
        # Total peso
        total_peso_cell = ws.cell(row=row, column=6, value=round(total_peso, 2))
        total_peso_cell.font = total_font
        total_peso_cell.fill = total_fill
        total_peso_cell.alignment = right_alignment

        # Ajustar ancho de columnas
        column_widths = [25, 30, 25, 12, 12, 15]  # Anchos específicos para cada columna
        for i, width in enumerate(column_widths, 1):
            column_letter = openpyxl.utils.get_column_letter(i)
            ws.column_dimensions[column_letter].width = width

        # Crear respuesta
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=consolidado_productos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return response

    except Exception as e:
        flash(f'Error al exportar: {str(e)}', 'error')
        return redirect(url_for('reportes.consolidado_productos'))
    finally:
        db.close()