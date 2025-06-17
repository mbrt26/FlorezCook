from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file
from config.database import db_config
from models import Cliente
from utils.helpers import get_current_year, DEPARTAMENTOS_CIUDADES
import pandas as pd
from sqlalchemy.exc import IntegrityError
import io
from datetime import datetime

clientes_bp = Blueprint('clientes', __name__, url_prefix='/clientes')  # Restore the prefix

# Move API route outside of the blueprint to handle it separately
api_clientes_bp = Blueprint('api_clientes', __name__, url_prefix='/api/clientes')

@api_clientes_bp.route('/buscar', methods=['GET'])
def buscar_api():
    """API para buscar cliente por NIT"""
    nit = request.args.get('nit')
    if not nit:
        return jsonify({'existe': False, 'error': 'NIT no proporcionado'}), 400
        
    db = db_config.get_session()
    try:
        cliente = db.query(Cliente).filter(Cliente.numero_identificacion == nit).first()
        if cliente:
            return jsonify({
                'existe': True,
                'id': cliente.id,
                'nombre_comercial': cliente.nombre_comercial,
                'razon_social': cliente.razon_social,
                'tipo_identificacion': cliente.tipo_identificacion,
                'numero_identificacion': cliente.numero_identificacion,
                'email': cliente.email,
                'telefono': cliente.telefono,
                'direccion': cliente.direccion,
                'ciudad': cliente.ciudad,
                'departamento': cliente.departamento
            })
        else:
            return jsonify({'existe': False})
    finally:
        db.close()

# Ruta específica para manejar /clientes sin barra final (evita redirect 308)
@clientes_bp.route('', methods=['GET'], strict_slashes=False)
@clientes_bp.route('/')
def lista():
    """Lista de clientes"""
    db = db_config.get_session()
    try:
        clientes = db.query(Cliente).all()
        current_year = get_current_year()
        return render_template('clientes_list.html', clientes=clientes, current_year=current_year)
    finally:
        db.close()

@clientes_bp.route('/agregar', methods=['GET', 'POST'])
def agregar():
    """Agregar nuevo cliente"""
    db = db_config.get_session()
    try:
        current_year = get_current_year()
        
        # Obtener parámetros de la URL para GET o del formulario para POST
        if request.method == 'POST':
            redirect_to = request.form.get('redirect_to', '')
            nit = request.form.get('nit', '')
        else:
            redirect_to = request.args.get('redirect_to', '')
            nit = request.args.get('nit', '')

        if request.method == 'POST':
            nombre = request.form.get('nombre_comercial')
            razon = request.form.get('razon_social')
            tipo = request.form.get('tipo_identificacion')
            numero = request.form.get('numero_identificacion')
            email = request.form.get('email')
            telefono = request.form.get('telefono')
            direccion = request.form.get('direccion')
            ciudad = request.form.get('ciudad')
            departamento = request.form.get('departamento')

            # Validación básica
            if not all([nombre, razon, tipo, numero, email, telefono, direccion, ciudad, departamento]):
                flash('Todos los campos son requeridos.', 'danger')
                return render_template('cliente_form.html', 
                                    current_year=current_year, 
                                    modo='agregar',
                                    nit_default=nit, 
                                    redirect_to=redirect_to,
                                    departamentos_ciudades=DEPARTAMENTOS_CIUDADES)
            
            # Verificar si ya existe un cliente con este número de identificación
            cliente_existente = db.query(Cliente).filter(Cliente.numero_identificacion == numero).first()
            if cliente_existente:
                flash(f'Ya existe un cliente registrado con el número de identificación {numero}. Por favor verifique el número o busque el cliente existente.', 'warning')
                return render_template('cliente_form.html', 
                                    current_year=current_year, 
                                    modo='agregar',
                                    nit_default=nit, 
                                    redirect_to=redirect_to,
                                    departamentos_ciudades=DEPARTAMENTOS_CIUDADES)
            
            try:
                # Crear el nuevo cliente
                cli = Cliente(
                    nombre_comercial=nombre,
                    razon_social=razon,
                    tipo_identificacion=tipo,
                    numero_identificacion=numero,
                    email=email,
                    telefono=telefono,
                    direccion=direccion,
                    ciudad=ciudad,
                    departamento=departamento
                )
                db.add(cli)
                db.commit()
                
                flash('Cliente agregado correctamente.', 'success')
                
                # Redirigir según corresponda
                if redirect_to:
                    if redirect_to == 'pedidos':
                        return redirect(url_for('pedidos.form', cliente_id=cli.id, show_welcome='true'))
                    elif redirect_to == '/pedidos/form':
                        return redirect('/pedidos/form')
                return redirect(url_for('clientes.lista'))
                
            except IntegrityError as e:
                db.rollback()
                # Manejar específicamente el error de clave duplicada
                if 'numero_identificacion' in str(e) or 'UNIQUE constraint failed' in str(e):
                    flash(f'Ya existe un cliente con el número de identificación {numero}. Por favor verifique el número.', 'warning')
                else:
                    flash('Error al guardar el cliente. Por favor intente nuevamente.', 'danger')
            except Exception as e:
                db.rollback()
                flash('Ocurrió un error inesperado. Por favor intente nuevamente.', 'danger')
                
        # Renderizar el formulario (GET o si hay error en POST)
        return render_template('cliente_form.html', 
                             current_year=current_year, 
                             modo='agregar',
                             nit_default=nit, 
                             redirect_to=redirect_to,
                             departamentos_ciudades=DEPARTAMENTOS_CIUDADES)
    finally:
        db.close()

@clientes_bp.route('/editar/<int:cliente_id>', methods=['GET', 'POST'])
def editar(cliente_id):
    """Editar cliente existente"""
    db = db_config.get_session()
    try:
        cli = db.query(Cliente).get(cliente_id)
        current_year = get_current_year()
        
        if not cli:
            flash('Cliente no encontrado.', 'danger')
            return redirect(url_for('clientes.lista'))
        
        if request.method == 'POST':
            nuevo_numero = request.form.get('numero_identificacion')
            
            # Verificar si se está cambiando el número de identificación y si ya existe
            if nuevo_numero != cli.numero_identificacion:
                cliente_existente = db.query(Cliente).filter(
                    Cliente.numero_identificacion == nuevo_numero,
                    Cliente.id != cliente_id
                ).first()
                
                if cliente_existente:
                    flash(f'Ya existe otro cliente con el número de identificación {nuevo_numero}. Por favor use un número diferente.', 'warning')
                    return render_template('cliente_form.html', 
                                         cliente=cli, 
                                         current_year=current_year, 
                                         modo='editar',
                                         departamentos_ciudades=DEPARTAMENTOS_CIUDADES)
            
            try:
                cli.nombre_comercial = request.form.get('nombre_comercial')
                cli.razon_social = request.form.get('razon_social')
                cli.tipo_identificacion = request.form.get('tipo_identificacion')
                cli.numero_identificacion = nuevo_numero
                cli.email = request.form.get('email')
                cli.telefono = request.form.get('telefono')
                cli.direccion = request.form.get('direccion')
                cli.ciudad = request.form.get('ciudad')
                cli.departamento = request.form.get('departamento')
                db.commit()
                flash('Cliente actualizado correctamente.', 'success')
                return redirect(url_for('clientes.lista'))
            except IntegrityError as e:
                db.rollback()
                # Manejar específicamente el error de clave duplicada
                if 'numero_identificacion' in str(e) or 'UNIQUE constraint failed' in str(e):
                    flash(f'Ya existe un cliente con el número de identificación {nuevo_numero}. Por favor verifique el número.', 'warning')
                else:
                    flash('Error al actualizar el cliente. Por favor intente nuevamente.', 'danger')
            except Exception as e:
                db.rollback()
                flash('Ocurrió un error inesperado. Por favor intente nuevamente.', 'danger')
        
        return render_template('cliente_form.html', 
                             cliente=cli, 
                             current_year=current_year, 
                             modo='editar',
                             departamentos_ciudades=DEPARTAMENTOS_CIUDADES)
    finally:
        db.close()

@clientes_bp.route('/eliminar/<int:cliente_id>', methods=['POST'])
def eliminar(cliente_id):
    """Eliminar cliente"""
    db = db_config.get_session()
    try:
        cli = db.query(Cliente).get(cliente_id)
        if cli:
            db.delete(cli)
            db.commit()
            flash('Cliente eliminado.', 'success')
        else:
            flash('Cliente no encontrado.', 'danger')
        return redirect(url_for('clientes.lista'))
    finally:
        db.close()

@clientes_bp.route('/ver/<int:cliente_id>')
def ver(cliente_id):
    """Ver detalles de un cliente"""
    db = db_config.get_session()
    try:
        cliente = db.query(Cliente).get(cliente_id)
        if not cliente:
            flash('Cliente no encontrado', 'danger')
            return redirect(url_for('clientes.lista'))
        return render_template('ver_cliente.html', cliente=cliente, current_year=get_current_year())
    finally:
        db.close()

@clientes_bp.route('/importar', methods=['GET', 'POST'])
def importar():
    """Importar clientes desde Excel"""
    db = db_config.get_session()
    try:
        current_year = get_current_year()
        
        if request.method == 'POST':
            if 'archivo' not in request.files:
                return redirect(request.url)
            
            archivo = request.files['archivo']
            if archivo.filename == '':
                return redirect(request.url)
                
            if archivo:
                try:
                    # Leer el archivo Excel
                    df = pd.read_excel(archivo)
                    
                    # Limpiar nombres de columnas (remover espacios extra y caracteres especiales)
                    df.columns = df.columns.str.strip().str.lower()
                    
                    # Mapeo de columnas permitidas (sin el asterisco para requeridas)
                    column_mapping = {
                        'nombre comercial': 'nombre_comercial',
                        'nombre comercial *': 'nombre_comercial',
                        'razón social': 'razon_social',
                        'razon social': 'razon_social',
                        'tipo identificación': 'tipo_identificacion',
                        'tipo identificacion': 'tipo_identificacion',
                        'número identificación': 'numero_identificacion',
                        'numero identificacion': 'numero_identificacion',
                        'email': 'email',
                        'teléfono': 'telefono',
                        'telefono': 'telefono',
                        'dirección': 'direccion',
                        'direccion': 'direccion',
                        'ciudad': 'ciudad',
                        'departamento': 'departamento'
                    }
                    
                    # Renombrar columnas según el mapeo
                    df = df.rename(columns=column_mapping)
                    
                    # Validar que al menos existe la columna requerida
                    if 'nombre_comercial' not in df.columns:
                        raise ValueError('La columna "Nombre Comercial" es obligatoria')
                    
                    # Procesar cada fila
                    resultados = {
                        'exito': True, 
                        'mensaje': 'Clientes importados correctamente', 
                        'errores': [],
                        'procesados': 0,
                        'exitosos': 0
                    }
                    
                    for index, fila in df.iterrows():
                        resultados['procesados'] += 1
                        fila_num = index + 2  # +2 porque pandas es 0-indexed y contamos el header
                        
                        try:
                            # Validar nombre comercial (campo obligatorio)
                            nombre_comercial = str(fila.get('nombre_comercial', '')).strip()
                            if not nombre_comercial or nombre_comercial.lower() in ['nan', 'none', '']:
                                resultados['errores'].append(f'Fila {fila_num}: El nombre comercial es obligatorio')
                                continue
                            
                            # Obtener y limpiar otros campos
                            numero_identificacion = str(fila.get('numero_identificacion', '')).strip()
                            if numero_identificacion.lower() in ['nan', 'none']:
                                numero_identificacion = ''
                            
                            # Si hay número de identificación, verificar duplicados
                            if numero_identificacion:
                                cliente_existente = db.query(Cliente).filter(
                                    Cliente.numero_identificacion == numero_identificacion
                                ).first()
                                if cliente_existente:
                                    resultados['errores'].append(
                                        f'Fila {fila_num}: Ya existe un cliente con el número de identificación {numero_identificacion}'
                                    )
                                    continue
                            
                            # Limpiar y preparar datos
                            def clean_field(value):
                                if pd.isna(value) or str(value).lower() in ['nan', 'none', '']:
                                    return ''
                                return str(value).strip()
                            
                            # Validar email si se proporciona
                            email = clean_field(fila.get('email', ''))
                            if email and '@' not in email:
                                resultados['errores'].append(f'Fila {fila_num}: Email inválido: {email}')
                                continue
                            
                            # Validar departamento si se proporciona
                            departamento = clean_field(fila.get('departamento', ''))
                            if departamento and departamento not in DEPARTAMENTOS_CIUDADES:
                                resultados['errores'].append(
                                    f'Fila {fila_num}: Departamento "{departamento}" no válido. '
                                    f'Use uno de: {", ".join(list(DEPARTAMENTOS_CIUDADES.keys())[:5])}...'
                                )
                                continue
                            
                            # Crear cliente
                            cliente = Cliente(
                                nombre_comercial=nombre_comercial,
                                razon_social=clean_field(fila.get('razon_social', '')) or nombre_comercial,
                                tipo_identificacion=clean_field(fila.get('tipo_identificacion', '')),
                                numero_identificacion=numero_identificacion,
                                email=email,
                                telefono=clean_field(fila.get('telefono', '')),
                                direccion=clean_field(fila.get('direccion', '')),
                                ciudad=clean_field(fila.get('ciudad', '')),
                                departamento=departamento
                            )
                            
                            db.add(cliente)
                            db.commit()
                            resultados['exitosos'] += 1
                            
                        except IntegrityError as e:
                            db.rollback()
                            if 'numero_identificacion' in str(e) or 'UNIQUE constraint failed' in str(e):
                                resultados['errores'].append(
                                    f'Fila {fila_num}: El número de identificación {numero_identificacion} ya está registrado'
                                )
                            else:
                                resultados['errores'].append(f'Fila {fila_num}: Error de base de datos - {str(e)}')
                        except Exception as e:
                            db.rollback()
                            resultados['errores'].append(f'Fila {fila_num}: {str(e)}')
                    
                    # Actualizar mensaje de resultado
                    if resultados['errores']:
                        resultados['exito'] = False
                        resultados['mensaje'] = f'Importación completada con errores. Procesados: {resultados["procesados"]}, Exitosos: {resultados["exitosos"]}, Errores: {len(resultados["errores"])}'
                    else:
                        resultados['mensaje'] = f'¡Importación exitosa! Se importaron {resultados["exitosos"]} clientes correctamente.'
                    
                    return render_template('importar_clientes.html', 
                                         resultados=resultados,
                                         current_year=current_year)
                    
                except Exception as e:
                    error_msg = f'Error al procesar el archivo: {str(e)}'
                    return render_template('importar_clientes.html', 
                                         resultados={'exito': False, 'mensaje': error_msg, 'errores': [error_msg]},
                                         current_year=current_year)
        
        return render_template('importar_clientes.html', current_year=current_year)
    finally:
        db.close()

@clientes_bp.route('/importar-clientes')
def redirect_importar():
    """Redirige a la página de importación de clientes"""
    return redirect(url_for('clientes.importar'))

@clientes_bp.route('/plantilla-excel')
def descargar_plantilla():
    """Descargar plantilla Excel para importar clientes"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.worksheet.datavalidation import DataValidation
        
        # Crear un nuevo workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Clientes"
        
        # Definir estilos
        header_fill = PatternFill(start_color="E85A0C", end_color="E85A0C", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        required_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        optional_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Encabezados de columna
        headers = [
            ('nombre_comercial', 'Nombre Comercial *', True),
            ('razon_social', 'Razón Social', False),
            ('tipo_identificacion', 'Tipo Identificación', False),
            ('numero_identificacion', 'Número Identificación', False),
            ('email', 'Email', False),
            ('telefono', 'Teléfono', False),
            ('direccion', 'Dirección', False),
            ('ciudad', 'Ciudad', False),
            ('departamento', 'Departamento', False)
        ]
        
        # Escribir encabezados
        for col, (field, header, required) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border
            
            # Ajustar ancho de columna
            column_letter = cell.column_letter
            ws.column_dimensions[column_letter].width = 20
        
        # Fila de ejemplo con datos reales
        ejemplo_data = [
            'Ejemplo Empresa S.A.S.',
            'Ejemplo Empresa S.A.S.',
            'NIT',
            '900123456-7',
            'contacto@ejemplo.com',
            '601 234 5678',
            'Carrera 15 # 93-47 Oficina 501',
            'Bogotá',
            'Cundinamarca'
        ]
        
        for col, value in enumerate(ejemplo_data, 1):
            cell = ws.cell(row=2, column=col, value=value)
            cell.border = border
            # Colorear celdas según si son requeridas u opcionales
            if headers[col-1][2]:  # Required field
                cell.fill = required_fill
            else:  # Optional field
                cell.fill = optional_fill
        
        # Agregar validación de datos para tipo_identificacion
        tipos_id = ['NIT', 'CC', 'CE', 'TI', 'RC', 'PA']
        dv_tipo_id = DataValidation(type="list", formula1='"{}"'.format(','.join(tipos_id)))
        dv_tipo_id.error = 'Tipo de identificación inválido'
        dv_tipo_id.errorTitle = 'Error de validación'
        ws.add_data_validation(dv_tipo_id)
        dv_tipo_id.add('C3:C1000')  # Aplicar a columna tipo_identificacion
        
        # Agregar validación para departamentos
        departamentos = list(DEPARTAMENTOS_CIUDADES.keys())
        dv_depto = DataValidation(type="list", formula1='"{}"'.format(','.join(departamentos)))
        dv_depto.error = 'Departamento inválido'
        dv_depto.errorTitle = 'Error de validación'
        ws.add_data_validation(dv_depto)
        dv_depto.add('I3:I1000')  # Aplicar a columna departamento
        
        # Crear hoja de instrucciones
        ws_inst = wb.create_sheet("Instrucciones")
        
        # Instrucciones detalladas
        instrucciones = [
            ["FlorezCook - Plantilla de Importación de Clientes", ""],
            ["", ""],
            ["INSTRUCCIONES DE USO:", ""],
            ["", ""],
            ["1. Campos Obligatorios (fondo rojo):", ""],
            ["   • nombre_comercial: Nombre comercial de la empresa", ""],
            ["", ""],
            ["2. Campos Opcionales (fondo azul):", ""],
            ["   • razon_social: Razón social (si es diferente al nombre comercial)", ""],
            ["   • tipo_identificacion: NIT, CC, CE, TI, RC, PA", ""],
            ["   • numero_identificacion: Número del documento", ""],
            ["   • email: Correo electrónico", ""],
            ["   • telefono: Número de teléfono", ""],
            ["   • direccion: Dirección completa", ""],
            ["   • ciudad: Ciudad donde se ubica", ""],
            ["   • departamento: Departamento (use la lista desplegable)", ""],
            ["", ""],
            ["3. Notas importantes:", ""],
            ["   • El número de identificación debe ser único", ""],
            ["   • Use el formato completo para teléfonos (ej: 601 234 5678)", ""],
            ["   • Para NIT incluya el dígito verificador (ej: 900123456-7)", ""],
            ["   • No modifique los encabezados de las columnas", ""],
            ["   • Elimine la fila de ejemplo antes de importar", ""],
            ["", ""],
            ["4. Departamentos y ciudades válidos:", ""]
        ]
        
        # Escribir instrucciones
        for row, (inst, desc) in enumerate(instrucciones, 1):
            ws_inst.cell(row=row, column=1, value=inst)
            if desc:
                ws_inst.cell(row=row, column=2, value=desc)
        
        # Formatear título de instrucciones
        title_cell = ws_inst.cell(row=1, column=1)
        title_cell.font = Font(bold=True, size=14, color="E85A0C")
        
        # Agregar lista de departamentos y ciudades
        current_row = len(instrucciones) + 2
        for depto, ciudades in DEPARTAMENTOS_CIUDADES.items():
            ws_inst.cell(row=current_row, column=1, value=depto)
            ws_inst.cell(row=current_row, column=2, value=", ".join(ciudades[:5]) + ("..." if len(ciudades) > 5 else ""))
            current_row += 1
        
        # Ajustar ancho de columnas en instrucciones
        ws_inst.column_dimensions['A'].width = 30
        ws_inst.column_dimensions['B'].width = 50
        
        # Guardar en buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generar nombre del archivo con fecha
        fecha_actual = datetime.now().strftime('%Y%m%d_%H%M%S')
        nombre_archivo = f'plantilla_clientes_florez_{fecha_actual}.xlsx'
        
        return send_file(
            output,
            as_attachment=True,
            download_name=nombre_archivo,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f'Error al generar la plantilla: {str(e)}', 'danger')
        return redirect(url_for('clientes.importar'))